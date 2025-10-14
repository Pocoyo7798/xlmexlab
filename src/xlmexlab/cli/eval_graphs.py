import json
from typing import Any, Dict
from zs4procext.parser import KeywordSearching

import click
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from zs4procext.evaluator_graphs import Evaluator_Graphs
import os


@click.command()
@click.argument('reference_file', type=click.Path(exists=True))
@click.argument('test_file', type=click.Path(exists=True))
@click.argument('output_file', type=click.Path())
@click.option('--label-threshold', default=0.7, help='Threshold for matching labels.')
@click.option('--series-threshold', default=0.7, help='Threshold for matching series.')
@click.option('--point-distance-threshold', default=0.1, help='Threshold for matching points based on euclidean distance.')

def main(reference_file: str, test_file: str, output_file: str, label_threshold: float, series_threshold: float, point_distance_threshold: float):
    """
    REFERENCE_FILE: Path to the reference JSON file.
    TEST_FILE: Path to the test JSON file.
    OUTPUT_FILE: Path to the output Excel file.
    """
    eval_graphs(reference_file, test_file, output_file, label_threshold, series_threshold, point_distance_threshold)

def eval_graphs(
    reference_file: str,
    test_file: str,
    output_file: str,
    label_threshold: float,
    series_threshold: float,
    point_distance_threshold: float
) -> None:
    reference_data = Evaluator_Graphs.load_json(reference_file)
    test_data = Evaluator_Graphs.load_json(test_file)
    
    evaluator = Evaluator_Graphs(reference_data=reference_data, threshold=label_threshold, distance_threshold=point_distance_threshold)
    
    # Process plots and get overall results
    results = evaluator.process_plots(test_data)

    # Calculate metrics for labels, series, and points
    label_metrics = evaluator.evaluate(results["Label_TP"], results["Label_FP"], results["Label_FN"])
    series_metrics = evaluator.evaluate(results["Series_TP"], results["Series_FP"], results["Series_FN"])
    point_metrics = evaluator.evaluate(results["Point_TP"], results["Point_FP"], results["Point_FN"])
    skipped_images = evaluator.evaluate (results["Skipped_Images"], results["Total_Images"], results["Skipped_Percent"])

    # Restructure the DataFrame
    data = {
        "Metric": ["FN", "FP", "TP", "Correct detections over Should have been Predicted", "Correct detections over what Was Predicted", "Correct Detections"],
        "Label Metrics": [results["Label_FN"], results["Label_FP"], results["Label_TP"], label_metrics["Correct detections over Should have been Predicted"], label_metrics["Correct detections over what Was Predicted"], label_metrics["Correct Detections"]],
        "Series Metrics": [results["Series_FN"], results["Series_FP"], results["Series_TP"], series_metrics["Correct detections over Should have been Predicted"], series_metrics["Correct detections over what Was Predicted"], series_metrics["Correct Detections"]],
        "Point Metrics": [results["Point_FN"], results["Point_FP"], results["Point_TP"], point_metrics["Correct detections over Should have been Predicted"], point_metrics["Correct detections over what Was Predicted"], point_metrics["Correct Detections"]],
        "Label Threshold": [label_threshold, "", "", "", "", ""],
        "Series Threshold": [series_threshold, "", "", "", "", ""],
        "Point Distance Threshold": [point_distance_threshold, "", "", "", "", ""],
        "Images Skipped":[results ["Skipped_Images"], "", "", "", "", ""],
        "Total Images": [results ["Total_Images"], "", "", "", "", ""],
        "% Images Skipped": [results ["Skipped_Percent"], "", "", "", "", ""]
    }

    df = pd.DataFrame(data)
    # Ensure the output file has a proper extension and filename
    if os.path.isdir(output_file):
        output_file = os.path.join(output_file, 'results.xlsx')
    elif not output_file.endswith(('.xlsx', '.xls')):
        output_file += '.xlsx'
    
    df.to_excel(output_file, index=False)

    # Apply color and bold formatting to specific rows in the Excel file
    wb = load_workbook(output_file)
    ws = wb.active
    
    bold_font = Font(bold=True)
    
    for row in [4, 5, 6]:  # Rows to bold (Correct detections over Should have been Predicted, Correct detections over what Was Predicted, Correct Detections)
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row + 1, column=col)  # +1 because Excel is 1-indexed
            cell.font = bold_font
    
    # Apply gradient color to numeric cells
    for row in range(5, ws.max_row + 1):
        for col in range(2, 5):  # Columns with numeric values
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                if cell.value >= 0.80:
                    fill = PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid")  # Pastel red
                elif cell.value >= 0.70:
                    fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Light red
                elif cell.value >= 0.45:
                    fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Light yellow
                elif cell.value >= 0.10:
                    fill = PatternFill(start_color="D4E9D6", end_color="D4E9D6", fill_type="solid")  # Lighter green
                else:
                    fill = PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid")  # Pastel green
                cell.fill = fill
    
    # Save the workbook
    wb.save(output_file)

    # Write per-image JSON
    json_output_file = output_file.replace(".xlsx", "_points.json")
    with open(json_output_file, "w") as f:
        json.dump(results["Per_Image_Results"], f, indent=4)
    print(f"Per-image point results saved to {json_output_file}")

    """"
    # Save matched label and series names with similarity
    match_file = output_file.replace('.xlsx', '_matches.xlsx')

    label_match_rows = [
        {"Type": "Label", "Plot": plot, "Reference": ref, "Test": test, "Similarity": ratio}
        for plot, ref, test, ratio in results.get("Matched_Labels", [])
    ]
    series_match_rows = [
        {"Type": "Series", "Plot": plot, "Reference": ref, "Test": test, "Similarity": ratio}
        for plot, ref, test, ratio in results.get("Matched_Series", [])
    ]

    match_df = pd.DataFrame(label_match_rows + series_match_rows)
    match_df.to_excel(match_file, index=False)
    print(f"\nMatch info saved to: {match_file}")

    # Save matched label and series names with similarity
    match_file = output_file.replace('.xlsx', '_matches.xlsx')

    label_match_rows = [
        {"Kind": "label", "Type": "Matched", "Plot": plot, "Reference": ref, "Test": test, "Similarity": ratio}
        for plot, ref, test, ratio in results.get("Matched_Labels", [])
    ]

    series_match_rows = [
        {"Kind": "series", "Type": "Matched", "Plot": plot, "Reference": ref, "Test": test, "Similarity": ratio}
        for plot, ref, test, ratio in results.get("Matched_Series", [])
    ]

    unmatched_items = results.get("Unmatched_Items", [])
    unmatched_rows = [
        {
            "Kind": kind,
            "Type": "Unmatched",
            "Plot": plot,
            "Reference": ref if source == "ref" else "",
            "Test": ref if source == "test" else "",
            "Similarity": ""  # No similarity computed
        }
        for plot, ref, source, kind in unmatched_items
    ]

    all_rows = label_match_rows + series_match_rows + unmatched_rows
    match_df = pd.DataFrame(all_rows)

    # Save to Excel
    match_df.to_excel(match_file, index=False)
    print(f"\nMatch info (matched + unmatched) saved to: {match_file}")

    """
        # ---- PER-IMAGE METRICS ----
    per_image_rows = []

    for plot_name in reference_data.keys():
        test_plot = test_data.get(plot_name)
        if not test_plot:
            per_image_rows.append({
                "Image": plot_name,
                #"Skipped": True,
                #"Label_TP": 0, "Label_FP": 0, "Label_FN": 0,
                #"Label_ND": "", "Label_MD": "", 
                "Label_ID": "",
                #"Series_TP": 0, "Series_FP": 0, "Series_FN": 0,
                #"Series_ND": "", "Series_MD": "", 
                "Series_ID": "",
                #"Point_TP": 0, "Point_FP": 0, "Point_FN": 0,
                #"Point_ND": "", "Point_MD": "", 
                "Point_ID": ""
            })
            continue

        ref_plot = reference_data[plot_name]

        ref_labels = evaluator.extract_labels({plot_name: ref_plot})
        test_labels = evaluator.extract_labels({plot_name: test_plot})
        label_TP, label_FP, label_FN, _, _, _ = evaluator.match_references_tests(ref_labels, test_labels)
        label_metrics = evaluator.evaluate(label_TP, label_FP, label_FN)

        ref_series = evaluator.extract_series({plot_name: ref_plot})
        test_series = evaluator.extract_series({plot_name: test_plot})
        series_TP, series_FP, series_FN, series_matches, matched_ref_series, matched_test_series = evaluator.match_references_tests(ref_series, test_series)
        series_metrics = evaluator.evaluate(series_TP, series_FP, series_FN)

        point_TP, point_FP, point_FN = evaluator.point_matching_accuracy(
            {plot_name: ref_plot},
            {plot_name: test_plot},
            series_matches,
            matched_ref_series,
            matched_test_series
        )
        point_metrics = evaluator.evaluate(point_TP, point_FP, point_FN)

        per_image_rows.append({
            "Image": plot_name,
            "Skipped": False,
            "Label_TP": label_TP, "Label_FP": label_FP, "Label_FN": label_FN,
            "Label_CoSP": label_metrics["Correct detections over Should have been Predicted"],
            "Label_CoWP": label_metrics["Correct detections over what Was Predicted"],
            "Label_CD": label_metrics["Correct Detections"],
            "Series_TP": series_TP, "Series_FP": series_FP, "Series_FN": series_FN,
            "Series_CoSP": series_metrics["Correct detections over Should have been Predicted"],
            "Series_CoWP": series_metrics["Correct detections over what Was Predicted"],
            "Series_CD": series_metrics["Correct Detections"],
            "Point_TP": point_TP, "Point_FP": point_FP, "Point_FN": point_FN,
            "Point_CoSP": point_metrics["Correct detections over Should have been Predicted"],
            "Point_CoWP": point_metrics["Correct detections over what Was Predicted"],
            "Point_CD": point_metrics["Correct Detections"]
        })

    per_image_df = pd.DataFrame(per_image_rows)
    per_image_file = output_file.replace(".xlsx", "_per_image_metrics.xlsx")
    per_image_df.to_excel(per_image_file, index=False)
    print(f"\nPer-image metrics saved to: {per_image_file}")


# ---- COMBINED CORRECT DETECTION SUMMARY FILE -------------
    summary_file = output_file.replace(".xlsx", "_accuracy_summary.xlsx")

    # Calculate percentage of successfully parsed images
    percent_success = 100 - results["Skipped_Percent"]

    # Extract the test file name (without full path)
    test_filename = os.path.basename(test_file)

    accuracy_data = {
        "File": [test_filename],
        "Correct Label Detection": [label_metrics["Correct Detections"]],
        "Correct Series Detection": [series_metrics["Correct Detections"]],
        "Correct Point Detection": [point_metrics["Correct Detections"]],
        "% Successfully Parsed Images": [percent_success]
    }

    accuracy_df = pd.DataFrame(accuracy_data)
    accuracy_df.to_excel(summary_file, index=False)

    print(f"\nCombined accuracy summary saved to: {summary_file}")





if __name__ == "__main__":
    main()
